from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from accounts.models import User

from .excel import QURILMA_USTUNLAR, XODIM_USTUNLAR
from .models import Bolim, Qurilma, Xodim


def excel_fayl(ustunlar, qatorlar, nom='test.xlsx'):
    """Test uchun xotirada .xlsx fayl yaratadi."""
    wb = Workbook()
    ws = wb.active
    ws.append([sarlavha for _, sarlavha, _ in ustunlar])
    for qator in qatorlar:
        ws.append(qator)
    bufer = BytesIO()
    wb.save(bufer)
    return SimpleUploadedFile(
        nom, bufer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


class AsosTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.bolim = Bolim.objects.create(nomi='Kutubxona')
        cls.boshqa_bolim = Bolim.objects.create(nomi='Dekanat')

        cls.userlar = {}
        for rol in User.Rol.values:
            cls.userlar[rol] = User.objects.create_user(username=f'u_{rol}', rol=rol)

        cls.xodim = Xodim.objects.create(
            user=cls.userlar[User.Rol.XODIM], fio='Test Xodim', bolim=cls.bolim,
        )
        cls.qurilma = Qurilma.objects.create(
            inventar_raqami='UNI-0001', turi=Qurilma.Turi.KOMPYUTER,
            bolim=cls.bolim, xodim=cls.xodim,
        )
        cls.begona = Qurilma.objects.create(
            inventar_raqami='UNI-0002', turi=Qurilma.Turi.PRINTER, bolim=cls.boshqa_bolim,
        )

    def kir(self, rol):
        self.client.force_login(self.userlar[rol])


class RuxsatTest(AsosTest):
    def test_royxatga_kirish_rollari(self):
        url = reverse('qurilma_royxati')
        for rol in [User.Rol.ADMIN, User.Rol.OPERATOR, User.Rol.USTA, User.Rol.XODIM]:
            with self.subTest(rol=rol):
                self.kir(rol)
                self.assertEqual(self.client.get(url).status_code, 200)

        self.kir(User.Rol.RAHBARIYAT)
        self.assertEqual(self.client.get(url).status_code, 403)

    def test_xodim_faqat_oz_qurilmalarini_koradi(self):
        self.kir(User.Rol.XODIM)
        javob = self.client.get(reverse('qurilma_royxati'))
        self.assertContains(javob, 'UNI-0001')
        self.assertNotContains(javob, 'UNI-0002')

    def test_xodim_begona_qurilma_sahifasiga_kira_olmaydi(self):
        self.kir(User.Rol.XODIM)
        self.assertEqual(
            self.client.get(reverse('qurilma_batafsil', args=[self.begona.pk])).status_code, 403,
        )
        self.assertEqual(
            self.client.get(reverse('qurilma_batafsil', args=[self.qurilma.pk])).status_code, 200,
        )

    def test_xodim_begona_qurilma_qr_kodini_ola_olmaydi(self):
        self.kir(User.Rol.XODIM)
        self.assertEqual(
            self.client.get(reverse('qurilma_qr', args=[self.begona.pk])).status_code, 403,
        )
        self.assertEqual(
            self.client.get(reverse('qurilma_qr', args=[self.qurilma.pk])).status_code, 200,
        )

    def test_qurilma_qoshish_faqat_admin_va_operator(self):
        url = reverse('qurilma_yangi')
        for rol in [User.Rol.USTA, User.Rol.XODIM, User.Rol.RAHBARIYAT]:
            with self.subTest(rol=rol):
                self.kir(rol)
                self.assertEqual(self.client.get(url).status_code, 403)

        self.kir(User.Rol.OPERATOR)
        self.assertEqual(self.client.get(url).status_code, 200)

    def test_import_faqat_admin_uchun(self):
        self.kir(User.Rol.OPERATOR)
        self.assertEqual(self.client.get(reverse('qurilma_import')).status_code, 403)
        self.kir(User.Rol.ADMIN)
        self.assertEqual(self.client.get(reverse('qurilma_import')).status_code, 200)

    def test_bolim_va_xodim_royxati_faqat_admin(self):
        for url_nomi in ['bolim_royxati', 'xodim_royxati']:
            with self.subTest(url=url_nomi):
                self.kir(User.Rol.OPERATOR)
                self.assertEqual(self.client.get(reverse(url_nomi)).status_code, 403)
                self.kir(User.Rol.ADMIN)
                self.assertEqual(self.client.get(reverse(url_nomi)).status_code, 200)


class QurilmaAmallarTest(AsosTest):
    def test_qurilma_qoshish(self):
        self.kir(User.Rol.OPERATOR)
        javob = self.client.post(reverse('qurilma_yangi'), {
            'inventar_raqami': 'UNI-0100',
            'turi': Qurilma.Turi.MONITOR,
            'modeli': 'LG 22MK430',
            'seriya_raqami': '',
            'bolim': self.bolim.pk,
            'xodim': '',
            'holati': Qurilma.Holat.ISHLAMOQDA,
            'olingan_sana': '',
        })
        yangi = Qurilma.objects.get(inventar_raqami='UNI-0100')
        self.assertRedirects(javob, yangi.get_absolute_url())
        self.assertEqual(yangi.modeli, 'LG 22MK430')

    def test_takroriy_inventar_raqami_qabul_qilinmaydi(self):
        self.kir(User.Rol.OPERATOR)
        javob = self.client.post(reverse('qurilma_yangi'), {
            'inventar_raqami': 'UNI-0001',
            'turi': Qurilma.Turi.MONITOR,
            'bolim': self.bolim.pk,
            'holati': Qurilma.Holat.ISHLAMOQDA,
        })
        self.assertEqual(javob.status_code, 200)
        self.assertEqual(Qurilma.objects.filter(inventar_raqami='UNI-0001').count(), 1)

    def test_yaroqsizga_chiqarish(self):
        self.kir(User.Rol.OPERATOR)
        self.client.post(
            reverse('qurilma_holat', args=[self.qurilma.pk]), {'holati': 'yaroqsiz'},
        )
        self.qurilma.refresh_from_db()
        self.assertEqual(self.qurilma.holati, Qurilma.Holat.YAROQSIZ)

    def test_qr_kod_rasmi(self):
        self.kir(User.Rol.OPERATOR)
        javob = self.client.get(reverse('qurilma_qr', args=[self.qurilma.pk]))
        self.assertEqual(javob['Content-Type'], 'image/png')
        self.assertTrue(javob.content.startswith(b'\x89PNG'))

    def test_qidiruv_va_filtr(self):
        self.kir(User.Rol.ADMIN)
        javob = self.client.get(reverse('qurilma_royxati'), {'q': 'UNI-0002'})
        self.assertContains(javob, 'UNI-0002')
        self.assertNotContains(javob, 'UNI-0001')

        javob = self.client.get(reverse('qurilma_royxati'), {'turi': Qurilma.Turi.PRINTER})
        self.assertContains(javob, 'UNI-0002')
        self.assertNotContains(javob, 'UNI-0001')


class ExcelTest(AsosTest):
    def setUp(self):
        self.kir(User.Rol.ADMIN)

    def test_shablon_yuklab_olinadi(self):
        javob = self.client.get(reverse('shablon', args=['qurilma']))
        self.assertEqual(javob.status_code, 200)
        wb = load_workbook(BytesIO(javob.content))
        sarlavhalar = [k.value for k in wb.active[1]]
        self.assertEqual(sarlavhalar, [s for _, s, _ in QURILMA_USTUNLAR])

    def test_qurilmalarni_import_qilish(self):
        fayl = excel_fayl(QURILMA_USTUNLAR, [
            ['UNI-0500', 'Printer', 'HP LaserJet', 'SN-1', 'Kutubxona', 'Test Xodim', '01.09.2024'],
            ['UNI-0501', 'kompyuter', 'Acer', '', 'Yangi bolim', '', ''],
        ])
        javob = self.client.post(reverse('qurilma_import'), {'fayl': fayl})
        natija = javob.context['natija']

        self.assertEqual(natija.yaratildi, 2)
        self.assertFalse(natija.xatolar)
        self.assertEqual(natija.yangi_bolimlar, ['Yangi bolim'])

        yangi = Qurilma.objects.get(inventar_raqami='UNI-0500')
        self.assertEqual(yangi.turi, Qurilma.Turi.PRINTER)
        self.assertEqual(yangi.xodim, self.xodim)
        self.assertEqual(str(yangi.olingan_sana), '2024-09-01')

    def test_import_xatolari_qator_raqami_bilan_korsatiladi(self):
        fayl = excel_fayl(QURILMA_USTUNLAR, [
            ['', 'Printer', '', '', 'Kutubxona', '', ''],                  # 2-qator: raqamsiz
            ['UNI-0600', 'Samolyot', '', '', 'Kutubxona', '', ''],         # 3-qator: turi xato
            ['UNI-0601', 'Printer', '', '', '', '', ''],                   # 4-qator: bo'limsiz
            ['UNI-0602', 'Printer', '', '', 'Kutubxona', '', 'kecha'],     # 5-qator: sana xato
            ['UNI-0603', 'Printer', '', '', 'Kutubxona', '', ''],          # 6-qator: to'g'ri
        ])
        natija = self.client.post(reverse('qurilma_import'), {'fayl': fayl}).context['natija']

        self.assertEqual([q for q, _ in natija.xatolar], [2, 3, 4, 5])
        self.assertEqual(natija.yaratildi, 1)
        self.assertTrue(Qurilma.objects.filter(inventar_raqami='UNI-0603').exists())

    def test_mavjud_qurilma_otkazib_yuboriladi_yoki_yangilanadi(self):
        qatorlar = [['UNI-0001', 'Monitor', 'Yangi model', '', 'Kutubxona', '', '']]

        natija = self.client.post(
            reverse('qurilma_import'), {'fayl': excel_fayl(QURILMA_USTUNLAR, qatorlar)},
        ).context['natija']
        self.assertEqual(natija.otkazildi, 1)
        self.qurilma.refresh_from_db()
        self.assertEqual(self.qurilma.turi, Qurilma.Turi.KOMPYUTER)

        natija = self.client.post(
            reverse('qurilma_import'),
            {'fayl': excel_fayl(QURILMA_USTUNLAR, qatorlar), 'yangilash': 'on'},
        ).context['natija']
        self.assertEqual(natija.yangilandi, 1)
        self.qurilma.refresh_from_db()
        self.assertEqual(self.qurilma.turi, Qurilma.Turi.MONITOR)
        self.assertEqual(self.qurilma.modeli, 'Yangi model')

    def test_notogri_sarlavhali_fayl_rad_etiladi(self):
        wb = Workbook()
        wb.active.append(['Nomi', 'Narxi'])
        bufer = BytesIO()
        wb.save(bufer)
        fayl = SimpleUploadedFile('xato.xlsx', bufer.getvalue())

        natija = self.client.post(reverse('qurilma_import'), {'fayl': fayl}).context['natija']
        self.assertEqual(len(natija.xatolar), 1)
        self.assertIn('sarlavhalari mos emas', natija.xatolar[0][1])

    def test_takrorlangan_inventar_raqami_faylda(self):
        fayl = excel_fayl(QURILMA_USTUNLAR, [
            ['UNI-0700', 'Printer', '', '', 'Kutubxona', '', ''],
            ['UNI-0700', 'Monitor', '', '', 'Kutubxona', '', ''],
        ])
        natija = self.client.post(reverse('qurilma_import'), {'fayl': fayl}).context['natija']
        self.assertEqual(natija.yaratildi, 1)
        self.assertIn('takrorlangan', natija.xatolar[0][1])

    def test_xodimlarni_import_qilish(self):
        fayl = excel_fayl(XODIM_USTUNLAR, [
            ['Aliyev Sardor', 'Kutubxona', 'Mutaxassis', '+998901234567', ''],
            ['Karimova Nodira', 'Dekanat', '', '', '123456789'],
        ])
        natija = self.client.post(reverse('xodim_import'), {'fayl': fayl}).context['natija']

        self.assertEqual(natija.yaratildi, 2)
        self.assertFalse(natija.xatolar)
        self.assertEqual(Xodim.objects.get(fio='Karimova Nodira').telegram_id, 123456789)

    def test_faqat_xlsx_qabul_qilinadi(self):
        fayl = SimpleUploadedFile('malumot.csv', b'a,b,c', content_type='text/csv')
        javob = self.client.post(reverse('qurilma_import'), {'fayl': fayl})
        self.assertIsNone(javob.context['natija'])
        self.assertFalse(javob.context['form'].is_valid())
