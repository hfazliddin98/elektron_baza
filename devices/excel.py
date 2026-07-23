"""Excel shablonlari va ommaviy import (qurilmalar, xodimlar)."""

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Bolim, Qurilma, Xodim

# (maydon, sarlavha, namuna qiymat)
QURILMA_USTUNLAR = [
    ('inventar_raqami', 'Inventar raqami*', 'UNI-0001'),
    ('turi', 'Turi*', 'Kompyuter'),
    ('modeli', 'Modeli', 'HP ProDesk 400 G7'),
    ('seriya_raqami', 'Seriya raqami', 'SN-123456'),
    ('bolim', "Bo'lim*", 'Kutubxona'),
    ('xodim', 'Xodim (F.I.Sh.)', 'Ismoilova Dilnoza'),
    ('olingan_sana', 'Olingan sana', '01.09.2024'),
]

XODIM_USTUNLAR = [
    ('fio', 'F.I.Sh.*', 'Yusupova Malika'),
    ('bolim', "Bo'lim*", 'Kutubxona'),
    ('lavozimi', 'Lavozimi', 'Kotiba'),
    ('telefon', 'Telefon', '+998901234567'),
    ('telegram_id', 'Telegram ID', ''),
]

USTUNLAR = {'qurilma': QURILMA_USTUNLAR, 'xodim': XODIM_USTUNLAR}
SARLAVHALAR = {'qurilma': 'Qurilmalar', 'xodim': 'Xodimlar'}


@dataclass
class ImportNatija:
    yaratildi: int = 0
    yangilandi: int = 0
    otkazildi: int = 0
    xatolar: list = field(default_factory=list)
    ogohlantirishlar: list = field(default_factory=list)
    yangi_bolimlar: list = field(default_factory=list)

    @property
    def muvaffaqiyatli(self):
        return not self.xatolar


def shablon_yarat(tur):
    """Import uchun namunaviy .xlsx fayl (bayt shaklida)."""
    ustunlar = USTUNLAR[tur]
    wb = Workbook()
    ws = wb.active
    ws.title = SARLAVHALAR[tur]

    sarlavha_uslubi = Font(bold=True, color='FFFFFF')
    fon = PatternFill('solid', fgColor='0D6EFD')

    for raqam, (_, sarlavha, namuna) in enumerate(ustunlar, start=1):
        katak = ws.cell(row=1, column=raqam, value=sarlavha)
        katak.font = sarlavha_uslubi
        katak.fill = fon
        katak.alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=raqam, value=namuna)
        ws.column_dimensions[get_column_letter(raqam)].width = max(len(sarlavha) + 4, 18)

    ws.freeze_panes = 'A2'
    izoh = ws.cell(row=4, column=1, value="* — majburiy maydon. 2-qatordagi namunani o'chirib, "
                                          "o'z ma'lumotlaringizni kiriting.")
    izoh.font = Font(italic=True, color='6C757D')

    bufer = BytesIO()
    wb.save(bufer)
    return bufer.getvalue()


def _matn(qiymat):
    if qiymat is None:
        return ''
    if isinstance(qiymat, float) and qiymat.is_integer():
        qiymat = int(qiymat)
    return str(qiymat).strip()


def _sana(qiymat):
    """Excel sanasi yoki '01.09.2024' / '2024-09-01' ko'rinishidagi matn."""
    if qiymat in (None, ''):
        return None
    if isinstance(qiymat, datetime):
        return qiymat.date()
    if isinstance(qiymat, date):
        return qiymat
    matn = _matn(qiymat)
    for format_ in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(matn, format_).date()
        except ValueError:
            continue
    raise ValueError(f"sana formati noto'g'ri: «{matn}» (masalan: 01.09.2024)")


def _turlar_xaritasi():
    xarita = {}
    for kod, nom in Qurilma.Turi.choices:
        xarita[kod.lower()] = kod
        xarita[nom.lower()] = kod
    return xarita


def _sarlavhalarni_tekshir(qator, ustunlar):
    kutilgan = [s.replace('*', '').strip().lower() for _, s, _ in ustunlar]
    bor = [_matn(k).replace('*', '').strip().lower() for k in (qator or [])]
    return bor[:len(kutilgan)] == kutilgan


def import_qil(fayl, tur, yangilash=False):
    """Excel faylni o'qib, bazaga yozadi. Natija va xatolarni qaytaradi."""
    natija = ImportNatija()
    ustunlar = USTUNLAR[tur]

    try:
        wb = load_workbook(fayl, data_only=True, read_only=True)
    except Exception as xato:
        natija.xatolar.append((0, f"Faylni o'qib bo'lmadi: {xato}"))
        return natija

    ws = wb.active
    qatorlar = list(ws.iter_rows(values_only=True))
    wb.close()

    if not qatorlar:
        natija.xatolar.append((0, "Fayl bo'sh."))
        return natija

    if not _sarlavhalarni_tekshir(qatorlar[0], ustunlar):
        kutilgan = ', '.join(s for _, s, _ in ustunlar)
        natija.xatolar.append((1, f'Fayl sarlavhalari mos emas. Kutilgan ustunlar: {kutilgan}. '
                                  'Shablonni yuklab olib, undan foydalaning.'))
        return natija

    bolimlar = {}

    def bolim_ol(nomi):
        kalit = nomi.lower()
        if kalit not in bolimlar:
            bolim, yaratildi = Bolim.objects.get_or_create(nomi=nomi)
            if yaratildi:
                natija.yangi_bolimlar.append(nomi)
            bolimlar[kalit] = bolim
        return bolimlar[kalit]

    turlar = _turlar_xaritasi()
    korilgan_inventarlar = set()

    for raqam, qator in enumerate(qatorlar[1:], start=2):
        qiymatlar = {maydon: _matn(qator[i]) if i < len(qator) else ''
                     for i, (maydon, _, _) in enumerate(ustunlar)}
        if not any(qiymatlar.values()):
            continue

        try:
            if tur == 'qurilma':
                _qurilma_qatori(qiymatlar, qator, raqam, natija, turlar, bolim_ol,
                                korilgan_inventarlar, yangilash, ustunlar)
            else:
                _xodim_qatori(qiymatlar, raqam, natija, bolim_ol, yangilash)
        except ValueError as xato:
            natija.xatolar.append((raqam, str(xato)))

    return natija


def _qurilma_qatori(q, xom_qator, raqam, natija, turlar, bolim_ol, korilganlar,
                    yangilash, ustunlar):
    inventar = q['inventar_raqami']
    if not inventar:
        raise ValueError('inventar raqami kiritilmagan')
    if inventar.lower() in korilganlar:
        raise ValueError(f'inventar raqami «{inventar}» faylda takrorlangan')
    korilganlar.add(inventar.lower())

    turi = turlar.get(q['turi'].lower())
    if not turi:
        nomlar = ', '.join(nom for _, nom in Qurilma.Turi.choices)
        raise ValueError(f'qurilma turi noto\'g\'ri: «{q["turi"]}» (mumkin: {nomlar})')

    if not q['bolim']:
        raise ValueError("bo'lim kiritilmagan")

    sana_ustuni = [m for m, _, _ in ustunlar].index('olingan_sana')
    sana = _sana(xom_qator[sana_ustuni] if sana_ustuni < len(xom_qator) else None)

    xodim = None
    if q['xodim']:
        xodim = Xodim.objects.filter(fio__iexact=q['xodim']).first()
        if xodim is None:
            natija.ogohlantirishlar.append(
                (raqam, f'«{q["xodim"]}» ismli xodim topilmadi — qurilma xodimsiz qo\'shildi')
            )

    mavjud = Qurilma.objects.filter(inventar_raqami__iexact=inventar).first()
    maydonlar = {
        'turi': turi, 'modeli': q['modeli'], 'seriya_raqami': q['seriya_raqami'],
        'bolim': bolim_ol(q['bolim']), 'xodim': xodim, 'olingan_sana': sana,
    }

    if mavjud:
        if not yangilash:
            natija.otkazildi += 1
            return
        for nom, qiymat in maydonlar.items():
            setattr(mavjud, nom, qiymat)
        mavjud.save()
        natija.yangilandi += 1
    else:
        Qurilma.objects.create(inventar_raqami=inventar, **maydonlar)
        natija.yaratildi += 1


def _xodim_qatori(q, raqam, natija, bolim_ol, yangilash):
    if not q['fio']:
        raise ValueError('F.I.Sh. kiritilmagan')
    if not q['bolim']:
        raise ValueError("bo'lim kiritilmagan")

    telegram_id = None
    if q['telegram_id']:
        try:
            telegram_id = int(q['telegram_id'])
        except ValueError:
            raise ValueError(f'Telegram ID raqam bo\'lishi kerak: «{q["telegram_id"]}»')

    bolim = bolim_ol(q['bolim'])
    mavjud = Xodim.objects.filter(fio__iexact=q['fio'], bolim=bolim).first()
    maydonlar = {'lavozimi': q['lavozimi'], 'telefon': q['telefon']}
    if telegram_id is not None:
        maydonlar['telegram_id'] = telegram_id

    if mavjud:
        if not yangilash:
            natija.otkazildi += 1
            return
        for nom, qiymat in maydonlar.items():
            setattr(mavjud, nom, qiymat)
        mavjud.save()
        natija.yangilandi += 1
    else:
        Xodim.objects.create(fio=q['fio'], bolim=bolim, **maydonlar)
        natija.yaratildi += 1
