"""Hisobotni Excel (.xlsx) faylga chiqarish."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SARLAVHA_FONT = Font(bold=True, color='FFFFFF')
SARLAVHA_FON = PatternFill('solid', fgColor='0D6EFD')


def _varaq(wb, nom, sarlavhalar, qatorlar, birinchimi=False):
    ws = wb.active if birinchimi else wb.create_sheet()
    ws.title = nom

    for raqam, sarlavha in enumerate(sarlavhalar, start=1):
        katak = ws.cell(row=1, column=raqam, value=sarlavha)
        katak.font = SARLAVHA_FONT
        katak.fill = SARLAVHA_FON
        katak.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(raqam)].width = max(len(str(sarlavha)) + 4, 16)

    for qator in qatorlar:
        ws.append(list(qator))

    ws.freeze_panes = 'A2'
    return ws


def hisobot_excel(malumot):
    """Hisobot ma'lumotidan ko'p varaqli .xlsx yaratadi."""
    davr = malumot['davr']
    umumiy = malumot['umumiy']
    wb = Workbook()

    # 1. Umumiy
    umumiy_qatorlar = [
        ('Davr', davr.nomi),
        ('Kelib tushgan murojaatlar', umumiy['kelgan']),
        ('Shundan shoshilinch', umumiy['shoshilinch']),
        ('Tugatilgan (topshirilgan)', umumiy['tugatilgan']),
        ('Rad etilgan', umumiy['rad_etilgan']),
        ('Hozir jarayonda', umumiy['jarayonda']),
        ('Muddati o\'tgan (SLA)', umumiy['kechikkan']),
        ('Qayta ta\'mirlar', umumiy['qayta_tamir']),
        ('O\'zi ta\'mirlagan (tasdiqlangan)', umumiy['ozi_tamir']),
        ('O\'zi ta\'mirlagan (rad etilgan)', umumiy['ozi_tamir_rad']),
        ('O\'rtacha ta\'mir muddati (kun)', umumiy['ortacha_muddat'] or '—'),
        ('O\'rtacha baho', round(umumiy['ortacha_baho'], 2) if umumiy['ortacha_baho'] else '—'),
        ('Baholanmagan', umumiy['baholanmagan']),
        ('Jami xarajat (so\'m)', float(umumiy['xarajat'])),
    ]
    umumiy_qatorlar += [('', '')] + [('Bosqich', "O'rtacha kun")]
    umumiy_qatorlar += [(nom, kun if kun is not None else '—')
                        for nom, kun in malumot['bosqichlar']]
    umumiy_qatorlar += [('', '')] + [('Murojaat manbasi', 'Soni')]
    umumiy_qatorlar += [(m['nomi'], m['soni']) for m in malumot['manbalar']]
    _varaq(wb, 'Umumiy', ['Ko\'rsatkich', 'Qiymat'], umumiy_qatorlar, birinchimi=True)

    # 2. Ustalar
    _varaq(wb, 'Ustalar',
           ['Usta', 'Mutaxassisligi', "Ta'mirlar", "O'rtacha baho",
            "Qayta ta'mir", "Qayta ta'mir %", 'Shaxsan so\'ralgan', 'Hozir aktiv'],
           [(u.fio, u.mutaxassisligi,
             u.tamirlar_soni,
             round(u.reyting, 2) if u.reyting else '—',
             u.qayta_soni, u.qayta_foiz, u.soralgan_soni, u.aktiv)
            for u in malumot['ustalar']])

    # 3. Xodimlar
    _varaq(wb, 'Xodimlar',
           ['Xodim', "Bo'limi", 'Murojaatlar', "O'zi ta'mirlagan"],
           [(x.fio, x.bolim.nomi, x.murojaat_soni, x.ozi_soni)
            for x in malumot['xodimlar']])

    # 4. Bo'limlar
    _varaq(wb, "Bo'limlar",
           ["Bo'lim", 'Qurilmalar', 'Murojaatlar'],
           [(b.nomi, b.qurilmalar_soni, b.murojaat_soni) for b in malumot['bolimlar']])

    # 5. Qurilmalar
    _varaq(wb, 'Qurilmalar',
           ['Inventar raqami', 'Turi', 'Modeli', "Bo'limi", "Ta'mirlar soni", 'Xarajat'],
           [(q.inventar_raqami, q.get_turi_display(), q.modeli, q.bolim.nomi,
             q.tamir_soni, float(q.xarajat or 0)) for q in malumot['qurilmalar']])

    # 6. Muddati o'tganlar
    _varaq(wb, "Muddati o'tganlar",
           ['#', 'Qurilma', 'Holat', 'Usta', 'Kechikish (kun)'],
           [(t.pk, t.qurilma.inventar_raqami, t.get_holat_display(),
             t.usta.fio if t.usta else '—', t.kechikkan_kun)
            for t in malumot['kechikkanlar']])

    bufer = BytesIO()
    wb.save(bufer)
    return bufer.getvalue()
