from datetime import date, timedelta
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import User, Usta
from devices.models import Bolim, Qurilma, Xodim
from repairs.models import TamirYozuvi

from .hisobot import Davr, yig

Holat = TamirYozuvi.Holat


class HisobotTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.bolim = Bolim.objects.create(nomi='Kutubxona')
        cls.admin = User.objects.create_user(username='a', rol=User.Rol.ADMIN)
        cls.rahbar = User.objects.create_user(username='r', rol=User.Rol.RAHBARIYAT)
        cls.operator = User.objects.create_user(username='o', rol=User.Rol.OPERATOR)

        usta_user = User.objects.create_user(username='u', rol=User.Rol.USTA)
        cls.usta = Usta.objects.create(user=usta_user, fio='Usta Bir')
        xodim_user = User.objects.create_user(username='x', rol=User.Rol.XODIM)
        cls.xodim = Xodim.objects.create(user=xodim_user, fio='Xodim Bir', bolim=cls.bolim)

        cls.qurilma = Qurilma.objects.create(
            inventar_raqami='UNI-0001', turi=Qurilma.Turi.PRINTER,
            bolim=cls.bolim, xodim=cls.xodim,
        )

        # Tugatilgan va baholangan ta'mir
        tamir = TamirYozuvi.objects.create(
            qurilma=cls.qurilma, xodim=cls.xodim, muammo_tavsifi='Test',
        )
        tamir.holat_ozgartir(Holat.QABUL, user=cls.operator)
        tamir.holat_ozgartir(Holat.TASHXIS, user=usta_user, usta=cls.usta)
        tamir.holat_ozgartir(Holat.TAMIRDA, user=usta_user)
        tamir.xarajat = 100000
        tamir.holat_ozgartir(Holat.TAYYOR, user=usta_user)
        tamir.holat_ozgartir(Holat.TOPSHIRILDI, user=cls.operator)
        TamirYozuvi.objects.filter(pk=tamir.pk).update(baho=5)
        cls.tamir = tamir

        # Tasdiqlangan «o'zim ta'mirladim»
        ozi = TamirYozuvi.objects.create(
            qurilma=cls.qurilma, xodim=cls.xodim, turi=TamirYozuvi.Turi.OZI,
            muammo_tavsifi='Kabel', bajarilgan_ishlar='Almashtirdim', xarajat=20000,
        )
        ozi.tasdiq_holati = TamirYozuvi.Tasdiq.TASDIQLANDI
        ozi.save()

        # Rad etilgan «o'zim ta'mirladim» — hisobga kirmasligi kerak
        rad = TamirYozuvi.objects.create(
            qurilma=cls.qurilma, xodim=cls.xodim, turi=TamirYozuvi.Turi.OZI,
            muammo_tavsifi='Test rad', bajarilgan_ishlar='...', xarajat=999999,
        )
        rad.tasdiq_holati = TamirYozuvi.Tasdiq.RAD
        rad.save()

    def davr(self):
        bugun = timezone.localdate()
        return Davr(bugun - timedelta(days=1), bugun + timedelta(days=1))

    def test_umumiy_korsatkichlar(self):
        malumot = yig(self.davr())
        umumiy = malumot['umumiy']

        self.assertEqual(umumiy['kelgan'], 1)
        self.assertEqual(umumiy['tugatilgan'], 1)
        self.assertEqual(umumiy['ozi_tamir'], 1)
        self.assertEqual(umumiy['ozi_tamir_rad'], 1)
        self.assertEqual(umumiy['ortacha_baho'], 5)
        # Rad etilgan o'zi-ta'mir xarajati hisobga kirmaydi
        self.assertEqual(float(umumiy['xarajat']), 120000)

    def test_ustalar_kesimi(self):
        usta = yig(self.davr())['ustalar'][0]
        self.assertEqual(usta.tamirlar_soni, 1)
        self.assertEqual(usta.reyting, 5)
        self.assertEqual(usta.qayta_soni, 0)

    def test_davr_tashqarisidagi_yozuv_hisobga_kirmaydi(self):
        otgan_davr = Davr(date(2020, 1, 1), date(2020, 12, 31))
        self.assertEqual(yig(otgan_davr)['umumiy']['kelgan'], 0)

    def test_davr_nomi(self):
        self.assertEqual(Davr(date(2026, 6, 1), date(2026, 6, 30)).nomi, '2026-yil iyun')
        self.assertEqual(Davr(date(2026, 1, 1), date(2026, 12, 31)).nomi, '2026-yil')
        self.assertEqual(Davr(date(2026, 6, 5), date(2026, 6, 20)).nomi,
                         '05.06.2026 — 20.06.2026')

    def test_sahifa_faqat_admin_va_rahbariyat_uchun(self):
        url = reverse('hisobot')
        self.client.force_login(self.operator)
        self.assertEqual(self.client.get(url).status_code, 403)

        for user in [self.admin, self.rahbar]:
            with self.subTest(user=user.username):
                self.client.force_login(user)
                self.assertEqual(self.client.get(url).status_code, 200)

    def test_excel_yuklab_olinadi(self):
        self.client.force_login(self.rahbar)
        javob = self.client.get(reverse('hisobot_excel'))
        self.assertEqual(javob.status_code, 200)

        wb = load_workbook(BytesIO(javob.content))
        self.assertIn('Ustalar', wb.sheetnames)
        self.assertIn('Umumiy', wb.sheetnames)
        self.assertEqual(wb['Ustalar']['A2'].value, 'Usta Bir')

    def test_pdf_yuklab_olinadi(self):
        self.client.force_login(self.rahbar)
        javob = self.client.get(reverse('hisobot_pdf'))
        self.assertEqual(javob.status_code, 200)
        self.assertEqual(javob['Content-Type'], 'application/pdf')
        self.assertTrue(javob.content.startswith(b'%PDF'))

    def test_davr_parametrlari_ishlaydi(self):
        self.client.force_login(self.admin)
        javob = self.client.get(reverse('hisobot'), {'dan': '2020-01-01', 'gacha': '2020-12-31'})
        self.assertEqual(javob.status_code, 200)
        self.assertEqual(javob.context['umumiy']['kelgan'], 0)

    def test_notogri_sana_joriy_oyga_qaytadi(self):
        self.client.force_login(self.admin)
        javob = self.client.get(reverse('hisobot'), {'dan': 'xato', 'gacha': '???'})
        self.assertEqual(javob.status_code, 200)
        self.assertEqual(javob.context['davr'].boshlanish.day, 1)
